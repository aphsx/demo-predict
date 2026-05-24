#!/usr/bin/env python3
"""
Import 1Moby Excel (8 sheets) into **train** raw tables (train_data_sources + train_raw_sheet_*).

Fidelity rules: trim header names only; store cell values as read (including Excel
date serials). No dedupe, no date coercion, no row drops except wholly empty rows.

Usage:
  export DATABASE_URL=postgresql://user:pass@localhost:5433/moby
  python scripts/import_train_raw.py \\
    --file "../data/[1Moby] Data_example for Bangkok university.xlsx" \\
    --name "Bangkok University example" \\
    --client bangkok_university
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import UUID

import psycopg
import yaml
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
SCHEMA_PATH = ROOT / "config" / "excel_schema.yaml"


def load_schema() -> dict[str, Any]:
    with SCHEMA_PATH.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def trim_header(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def cell_to_json(value: Any) -> Any:
    """Preserve values without business transforms (dates stay serial when applicable)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        # openpyxl may return datetime for date cells — store ISO + serial for audit
        serial = (value - datetime(1899, 12, 30)).total_seconds() / 86400
        return {"_excel": "datetime", "iso": value.isoformat(), "serial": serial}
    if isinstance(value, date):
        dt = datetime(value.year, value.month, value.day)
        serial = (dt - datetime(1899, 12, 30)).total_seconds() / 86400
        return {"_excel": "date", "iso": value.isoformat(), "serial": serial}
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def row_is_empty(values: list[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def build_payload(headers: list[str | None], row_values: tuple[Any, ...]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for i, key in enumerate(headers):
        if key is None:
            continue
        if i >= len(row_values):
            payload[key] = None
        else:
            payload[key] = cell_to_json(row_values[i])
    return payload


def validate_headers(sheet_name: str, headers_trimmed: list[str | None], sheet_cfg: dict) -> None:
    present = {h for h in headers_trimmed if h is not None}
    for req in sheet_cfg.get("required_headers", []):
        if req not in present:
            raise ValueError(
                f"Sheet {sheet_name!r}: missing required header {req!r}. "
                f"Found: {sorted(present)}"
            )


def iter_sheet_rows(ws: Worksheet, skip_empty: bool):
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], []

    headers = [trim_header(c) for c in header_row]
    out: list[tuple[int, dict[str, Any]]] = []
    excel_row = 1
    for row_values in rows:
        excel_row += 1
        if skip_empty and row_is_empty(list(row_values)):
            continue
        payload = build_payload(headers, row_values)
        out.append((excel_row, payload))
    return headers, out


def import_sheet(
    conn: psycopg.Connection,
    source_id: UUID,
    table: str,
    rows: list[tuple[int, dict[str, Any]]],
    batch_size: int,
) -> int:
    if not rows:
        return 0
    sql = f"""
        INSERT INTO {table} (source_id, excel_row, row_payload)
        VALUES (%s, %s, %s::jsonb)
    """
    count = 0
    with conn.cursor() as cur:
        batch: list[tuple] = []
        for excel_row, payload in rows:
            batch.append((source_id, excel_row, json.dumps(payload, ensure_ascii=False)))
            if len(batch) >= batch_size:
                cur.executemany(sql, batch)
                count += len(batch)
                batch = []
        if batch:
            cur.executemany(sql, batch)
            count += len(batch)
    return count


def main() -> int:
    parser = argparse.ArgumentParser(description="Import 1Moby Excel into raw tables")
    parser.add_argument("--file", required=True, type=Path, help="Path to .xlsx")
    parser.add_argument("--name", required=True, help="Human-readable source name")
    parser.add_argument("--client", default=None, help="client_label e.g. bangkok_university")
    parser.add_argument("--notes", default=None)
    parser.add_argument(
        "--schema",
        type=Path,
        default=SCHEMA_PATH,
        help="Path to excel_schema.yaml",
    )
    args = parser.parse_args()

    xlsx_path = args.file.resolve()
    if not xlsx_path.is_file():
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        return 1

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("DATABASE_URL is required", file=sys.stderr)
        return 1

    with args.schema.open(encoding="utf-8") as f:
        schema = yaml.safe_load(f)

    checksum = sha256_file(xlsx_path)
    file_size = xlsx_path.stat().st_size
    batch_size = int(schema.get("import", {}).get("batch_size", 500))
    skip_empty = bool(schema.get("import", {}).get("skip_wholly_empty_rows", True))

    sheet_map: dict[str, dict] = schema["sheets"]
    required = set(schema.get("required_sheets", []))
    optional = set(schema.get("optional_sheets", []))
    all_known = required | optional

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    present_sheets = set(wb.sheetnames)

    missing_required = required - present_sheets
    if missing_required:
        print(f"Missing required sheets: {sorted(missing_required)}", file=sys.stderr)
        wb.close()
        return 1

    unknown = present_sheets - all_known
    if unknown:
        print(f"Warning: unknown sheets (skipped): {sorted(unknown)}", file=sys.stderr)

    to_import = [s for s in wb.sheetnames if s in sheet_map]

    with psycopg.connect(db_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, import_status FROM train_data_sources WHERE file_checksum_sha256 = %s",
                (checksum,),
            )
            existing = cur.fetchone()
            if existing:
                print(
                    f"Already imported (checksum match): source_id={existing[0]} "
                    f"status={existing[1]}",
                    file=sys.stderr,
                )
                wb.close()
                return 2

            cur.execute(
                """
                INSERT INTO train_data_sources (
                    name, client_label, original_filename,
                    file_checksum_sha256, file_size_bytes, import_status
                ) VALUES (%s, %s, %s, %s, %s, 'importing')
                RETURNING id
                """,
                (args.name, args.client, xlsx_path.name, checksum, file_size),
            )
            source_id = cur.fetchone()[0]
        conn.commit()

        manifest: dict[str, int] = {}
        try:
            for sheet_name in to_import:
                cfg = sheet_map[sheet_name]
                table = cfg["table"]
                ws = wb[sheet_name]
                headers, rows = iter_sheet_rows(ws, skip_empty=skip_empty)
                validate_headers(sheet_name, headers, cfg)
                n = import_sheet(conn, source_id, table, rows, batch_size)
                manifest[sheet_name] = n
                print(f"  {sheet_name}: {n} rows → {table}")

            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE train_data_sources
                    SET import_status = 'ready',
                        imported_at = NOW(),
                        sheet_manifest = %s::jsonb,
                        notes = COALESCE(%s, notes)
                    WHERE id = %s
                    """,
                    (json.dumps(manifest), args.notes, source_id),
                )
            conn.commit()
            print(f"Done. source_id={source_id}")
            print(f"manifest={json.dumps(manifest, indent=2)}")
        except Exception as e:
            conn.rollback()
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM train_data_sources WHERE id = %s",
                    (source_id,),
                )
            conn.commit()
            print(f"Import failed (rolled back): {e}", file=sys.stderr)
            wb.close()
            return 1

    wb.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
